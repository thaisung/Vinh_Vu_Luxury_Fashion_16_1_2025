from ..models import *

from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_list_or_404, get_object_or_404
from django.core.paginator import Paginator


from django.http import HttpResponse
import requests
import time

from django.db import models
from django.utils import timezone

import os

from datetime import datetime

from django.shortcuts import redirect
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login, logout

from django.contrib.postgres.search import TrigramSimilarity
from django.db.models import Q
from django.shortcuts import render, redirect, reverse
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from datetime import datetime
from django.contrib import messages
import random
import string
from django.contrib.auth import update_session_auth_hash
from datetime import datetime, timedelta
from django.utils.timezone import make_aware

# from PIL import Image, ImageDraw, ImageFont
import requests
from io import BytesIO

import random
import string

import base64

import time
from django.http import JsonResponse

import re
import json

from django.conf import settings
from django.db.models import Q

import datetime

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_v1_5
import base64

from django.core.mail import send_mail,EmailMessage
from django.conf import settings
import openpyxl

from django.db.models import Count
from django.http import JsonResponse
from django.forms.models import model_to_dict

import tempfile

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.core.exceptions import ObjectDoesNotExist
import uuid
from django.db.models import Case, When, BooleanField
from django.utils.html import escape
import math


def product_ad(request,Slug):
    if request.method == 'GET':
        if request.user.is_authenticated:
            if request.user.is_superuser:
                context = {}
                s = request.GET.get('s')
                f = request.GET.get('f')
                st = request.GET.get('st')
                context = {}
                context['obj_Category_product_child'] = Category_product_child.objects.get(Slug=Slug)
                context['List_Category_product'] = Category_product.objects.all()
                context['List_Trademark'] = Trademark.objects.all()
                context['List_Product'] = context['obj_Category_product_child'].category_product_child_detail.all()
                if s:
                    context['List_Product'] = context['List_Product'].filter(Q(Name__icontains=s)).order_by('-id')
                    context['s'] = s
                if f:
                    Belong_Campaign = Product.objects.get(pk=f)
                    context['List_Product'] = context['List_Product'].filter(Belong_Campaign=Belong_Campaign).order_by('-id')
                    context['f'] = int(f)
                if st:
                    context['List_Product'] = context['List_Product'].filter(Status_lead=st).order_by('-id')
                    context['st'] = st               
                # Sử dụng Paginator để chia nhỏ danh sách (10 là số lượng mục trên mỗi trang)
                paginator = Paginator(context['List_Product'], settings.PAGE)

                # Lấy số trang hiện tại từ URL, nếu không mặc định là trang 1
                p = request.GET.get('p')
                page_obj = paginator.get_page(p)
                context['List_Product'] = page_obj
                # Tạo danh sách các số trang
                page_list = list(range(1, paginator.num_pages + 1))
                context['page_list'] = page_list
                return render(request, 'sleekweb/admin/product.html', context, status=200)
            else:
                return JsonResponse({'success': False, 'message': 'Bạn chưa được cấp quyền để thực hiện chức năng'},json_dumps_params={'ensure_ascii': False})
        else:
            return redirect('login_page_client')
    elif request.method == 'POST':
        if request.user.is_authenticated:
            if request.user.is_superuser:
                Slug = Slug
                fields={}
                fields['Name'] = request.POST.get('Name')
                if fields['Name']:
                    if Product.objects.filter(Name=fields['Name']):
                        return JsonResponse({'success': False, 'message': 'Tên Sản phẩm đã tồn tại trong hệ thống bao gồm cả Thư mục khác, vui lòng chọn tên khác để tạo Sản phẩm'},json_dumps_params={'ensure_ascii': False})
                else:
                    return JsonResponse({'success': False, 'message': 'Tên Sản phẩm không được để trống'},json_dumps_params={'ensure_ascii': False})
                fields['Slug'] = slugify(fields['Name'])
                fields['Price'] = request.POST.get('Price')
                fields['Description'] = request.POST.get('Description')
                fields['Discount'] = request.POST.get('Discount')
                if fields['Discount'] and fields['Price']:
                    fields['Price_Discount'] = math.ceil((100-int(fields['Discount']))/100*int(fields['Price']))
                else:
                    fields['Discount'] = 0
                Belong_Trademark = request.POST.get('Belong_Trademark')
                if Belong_Trademark:
                    fields['Belong_Trademark'] = Trademark.objects.get(pk=Belong_Trademark)
                fields['Belong_Category_product_child'] = Category_product_child.objects.get(Slug=Slug)
                obj_create_product = Product.objects.create(**fields)
                files_img  =  request.FILES.getlist('files_img')
                for i in files_img:
                    Photo_product.objects.create(Photo=i,Belong_Product_Photo=obj_create_product)
                return redirect('product_ad',Slug=Slug)
            else:
                return JsonResponse({'success': False, 'message': 'Bạn chưa được cấp quyền để thực hiện chức năng'},json_dumps_params={'ensure_ascii': False})
        else:
            return redirect('login_page_client')
    else:
        return redirect('login_page_client')
    
def update_product_ad(request,Slug,Slugg):
    if request.method == 'GET':
        obj_product=Product.objects.get(Slug=Slugg)
        list_Trademark = Trademark.objects.all()
        # Tạo danh sách các thẻ <option> cho thương hiệu
        trademark_options = '<option value="">--Không Thương hiệu--</option>'
        for trademark in list_Trademark:
            selected = 'selected' if obj_product.Belong_Trademark and obj_product.Belong_Trademark.id == trademark.id else ''
            trademark_options += f'<option value="{trademark.id}" {selected}>{escape(trademark.Name)}</option>'
        html_content=f"""
                    <label class="block text-sm w-full">
                        <span class="font-medium text-stone-600">Tên Sản phẩm</span>
                        <input name="Name" required value="{escape(obj_product.Name)}"
                            class="block w-full  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md">
                    </label>
                    <label class="block text-sm w-full">
                        <span class="font-medium text-stone-600">Giá </span>
                        <input name="Price" required value="{escape(obj_product.Price)}"  type="number" min="1000"
                            class="block w-full  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md">
                    </label>
                    <label class="flex flex-col text-sm w-full h-[150px] col-span-2">
                        <span class="font-medium text-stone-600">Mô tả</span>
                        <textarea name="Description" required  class="w-full grow block  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md ">{escape(obj_product.Description)}</textarea>
                    </label>
                    <label class="block text-sm w-full">
                        <span class="font-medium text-stone-600">Giảm giá (%)</span>
                        <input name="Discount" value="{escape(obj_product.Discount)}" type="number" min="0" max="100"
                            class="block w-full  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md ">
                    </label>
                    <label class="flex flex-col text-sm w-full">
                        <span class="font-medium text-stone-600">Thương hiệu</span>
                        <select name="Belong_Trademark" required
                        class="block w-full  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md ">
                            {trademark_options}
                        </select>
                    </label>
                    <label class="block text-sm w-full col-span-2">
                        <span class="font-medium text-stone-600">Ảnh sản phẩm mới</span>
                        <input name="files_img" type="file" multiple accept=".jpg,.png,.webp,.jpeg" onchange="sortFiles()"
                            class="block w-full  text-sm dark:border-gray-600 dark:bg-gray-700 focus:border-green-400 focus:outline-none focus:shadow-outline-green dark:text-gray-300 dark:focus:shadow-outline-gray form-input rounded-md ">
                    </label>
        """
        return HttpResponse(html_content, content_type="text/html")
    elif request.method == 'POST':
        if request.user.is_authenticated:
            if request.user.is_superuser:
                Slug = Slug
                Name = request.POST.get('Name')
                if Name:
                    obj_product = Product.objects.get(Slug=Slugg)
                    list_product = Product.objects.all().exclude(id=obj_product.id).filter(Name=Name)
                    if list_product:
                        return JsonResponse({'success': False, 'message': 'Tên Sản phẩm đã tồn tại trong hệ thống bao gồm cả Thư mục khác, vui lòng chọn tên khác để cập nhật Sản phẩm'},json_dumps_params={'ensure_ascii': False})
                else:
                    return JsonResponse({'success': False, 'message': 'Tên Sản phẩm không được để trống'},json_dumps_params={'ensure_ascii': False})
                obj_product.Name = Name
                obj_product.Slug = slugify(Name)
                obj_product.Price = request.POST.get('Price')
                obj_product.Description = request.POST.get('Description')
                obj_product.Discount = request.POST.get('Discount')
                if request.POST.get('Discount') and request.POST.get('Price'):
                    obj_product.Price_Discount = math.ceil((100-int(request.POST.get('Discount')))/100*int(request.POST.get('Price')))
                else:
                    obj_product.Price_Discount = 0
                Belong_Trademark = request.POST.get('Belong_Trademark')
                if Belong_Trademark:
                    obj_product.Belong_Trademark = Trademark.objects.get(pk=Belong_Trademark)
                obj_product.Belong_Category_product_child = Category_product_child.objects.get(Slug=Slug)
                obj_product.save()
                files_img  =  request.FILES.getlist('files_img')
                files_img.sort(key=lambda f: f.name)  # f.name là tên của file
                print('files_img:',files_img)
                if files_img:
                    Photo_product.objects.filter(Belong_Product_Photo=obj_product).delete()
                    for i in files_img:
                        Photo_product.objects.create(Photo=i,Belong_Product_Photo=obj_product)
                return redirect('product_ad',Slug=Slug)
            else:
                return JsonResponse({'success': False, 'message': 'Bạn chưa được cấp quyền để thực hiện chức năng'},json_dumps_params={'ensure_ascii': False})
        else:
            return redirect('login_page_client')
    else:
        return redirect('login_page_client')

def delete_check_list_product_ad(request,Slug):
    if request.method == 'POST':
        if request.user.is_authenticated:
            if request.user.is_superuser or request.user.is_manage or request.user.is_staff:
                data = json.loads(request.body)
                print('data:',data)
                check_list  = json.loads(data.get('check_list'))
                text_check_list = data.get('text_check_list')
                if text_check_list == 'Tôi đồng ý xóa danh sách Sản phẩm đã chọn':
                    Product.objects.filter(id__in=check_list).delete()
                    return JsonResponse({'success': True, 'redirect_url': reverse('product_ad', kwargs={'Slug': Slug})},json_dumps_params={'ensure_ascii': False})
                else:
                    return JsonResponse({'success': False, 'message': 'Nội dung nhập chưa chính xác'},json_dumps_params={'ensure_ascii': False})
            else:
                return JsonResponse({'success': False, 'message': 'Tài khoản của bạn chưa được cấp quyền để thực hiện chức năng này'},json_dumps_params={'ensure_ascii': False})
        else:
                return redirect('login_page_client')
    else:
        return redirect('product_ad',Slug=Slug)

def upload_file_lead_page_ad(request):
    if request.method == 'POST':
        if request.user.is_authenticated:
            if request.user.is_superuser or request.user.is_manage:
                file = request.FILES.get('file_upload')
                # Mở file bằng openpyxl (lưu ý file phải là định dạng Excel hợp lệ)
                workbook = openpyxl.load_workbook(file)

                # Duyệt qua tất cả các sheet trong file Excel
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    print(f"Đang xử lý sheet: {sheet}")

                    # Lặp qua các hàng bắt đầu từ hàng thứ 2 (bỏ qua tiêu đề hoặc hàng đầu tiên)
                    for row in worksheet.iter_rows(min_row=2, max_col=10, values_only=True):
                        field={}
                        field['Name'] = row[1]
                        field['Phone_number'] = return_phone_number(row[2])
                        field['Area'] = row[3]
                        field['Product'] = row[4]
                        field['Demand'] = row[5]
                        field['Status'] = 'Khách hàng mới'
                        field['Note'] = row[6]
                        field['Source'] = row[7]
                        
                        print('field:',field)
                        
                        Belong_Campaign_xlsx = row[0]
                        if request.user.is_superuser:
                            if Belong_Campaign_xlsx:
                                list_Campaign_xlsx = Product.objects.filter(Name=Belong_Campaign_xlsx)
                                if list_Campaign_xlsx:
                                    field['Belong_Campaign'] = list_Campaign_xlsx[0]
                                    try:
                                        Product.objects.get(Phone_number=field['Phone_number'],Belong_Campaign=field['Belong_Campaign'])
                                        return JsonResponse({'success': False, 'message': 'Số điện thoại trong chiến dịch đã tồn tại.'},json_dumps_params={'ensure_ascii': False})
                                    except:
                                        print('ok')
                                    field['Belong_User'] = list_Campaign_xlsx[0].Belong_User.all().annotate(lead_count=Count('List_Product')).exclude(is_superuser=True).exclude(is_manage=True).order_by('lead_count').first()
                                else:
                                    obj_Campaign = Product.objects.create(Name=Belong_Campaign_xlsx)
                                    field['Belong_Campaign'] = obj_Campaign
                            obj_lead = Product.objects.create(**field)
                            History.objects.create(history='Tạo mới khách hàng',Belong_Lead=obj_lead)
                        if request.user.is_manage:
                            if Belong_Campaign_xlsx:
                                list_Campaign_xlsx = Product.objects.filter(Name=Belong_Campaign_xlsx)
                                list_campaign = Product.objects.filter(Belong_User=request.user)
                                if list_Campaign_xlsx:
                                    obj_campaign =  list_Campaign_xlsx[0]
                                    if obj_campaign in list_campaign:
                                        field['Belong_Campaign'] = obj_campaign
                                        try:
                                            Product.objects.get(Phone_number=field['Phone_number'],Belong_Campaign=field['Belong_Campaign'])
                                            return JsonResponse({'success': False, 'message': 'Số điện thoại trong chiến dịch đã tồn tại.'},json_dumps_params={'ensure_ascii': False})
                                        except:
                                            print('ok')
                                        field['Belong_User'] = list_Campaign_xlsx[0].Belong_User.all().annotate(lead_count=Count('List_Product')).exclude(is_superuser=True).exclude(is_manage=True).order_by('lead_count').first()
                                    else:
                                        pass
                                else:
                                    obj_Campaign = Product.objects.create(Name=Belong_Campaign_xlsx)
                                    obj_Product.Belong_User.add(request.user)
                                    field['Belong_Campaign'] = obj_Campaign
                            obj_lead = Product.objects.create(**field)
                            History.objects.create(history='Tạo mới khách hàng',Belong_Lead=obj_lead)
                        if all(cell is None for cell in row):
                            break
                        # Xử lý dữ liệu trong hàng (từ cột A đến J)
                        print(row)
                if request.user.is_superuser or request.user.is_manage:  
                    return redirect('lead_page_admin')
            else:
                return JsonResponse({'success': False, 'message': 'Bạn chưa được cấp quyền để thực hiện chức năng'},json_dumps_params={'ensure_ascii': False})
        else:
            return redirect('login_page_client')
    else:
        return redirect('login_page_client')


